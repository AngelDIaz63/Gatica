"""
Script de demostración - Genera responsivas sin interfaz gráfica
Útil para probar la funcionalidad básica del sistema
"""

import sys
import os
from pathlib import Path
from datetime import datetime

# Importar solo las clases necesarias (sin tkinter)
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class DatabaseManager:
    """Gestiona la base de datos SQLite local"""
    
    def __init__(self, db_path='responsivas_demo.db'):
        self.db_path = db_path
        self.init_database()
    
    def get_connection(self):
        return sqlite3.connect(self.db_path)
    
    def init_database(self):
        """Inicializa la base de datos con datos de ejemplo"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Crear tablas
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS empresas (
                id_empresa INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL UNIQUE
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS areas (
                id_area INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL UNIQUE
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS usuarios (
                id_usuario INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                cedula TEXT,
                cargo TEXT,
                id_area INTEGER,
                id_empresa INTEGER,
                lap_marca TEXT,
                lap_modelo TEXT,
                lap_serie TEXT UNIQUE,
                lap_estado TEXT,
                sts TEXT DEFAULT 'ACTIVO',
                fecha_asignacion DATE
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS accesorios (
                id_accesorio INTEGER PRIMARY KEY AUTOINCREMENT,
                id_usuario INTEGER,
                tipo TEXT NOT NULL,
                marca TEXT,
                modelo TEXT,
                serie TEXT,
                estado TEXT DEFAULT 'BUENO'
            )
        ''')
        
        # Insertar datos de ejemplo
        cursor.execute("SELECT COUNT(*) FROM empresas")
        if cursor.fetchone()[0] == 0:
            cursor.execute("INSERT INTO empresas (nombre) VALUES (?)", ("VUBA LOGISTICS",))
            
            cursor.execute("INSERT INTO areas (nombre) VALUES (?)", ("Tecnología",))
            
            cursor.execute('''
                INSERT INTO usuarios (nombre, cedula, cargo, id_area, id_empresa, 
                                    lap_marca, lap_modelo, lap_serie, lap_estado, fecha_asignacion)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', ("Juan Pérez García", "12345678", "Analista TI", 1, 1, 
                  "DELL", "Latitude 5420", "SN-DELL-2024-001", "EXCELENTE", "2024-01-15"))
            
            cursor.execute('''
                INSERT INTO accesorios (id_usuario, tipo, marca, modelo, serie, estado)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (1, "MOUSE", "Logitech", "M720", "SER-MOUSE-001", "BUENO"))
            
            cursor.execute('''
                INSERT INTO accesorios (id_usuario, tipo, marca, modelo, serie, estado)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (1, "TECLADO", "Logitech", "K380", "SER-TECLADO-001", "BUENO"))
        
        conn.commit()
        conn.close()
    
    def buscar_usuario(self, nombre):
        """Busca un usuario por nombre"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT u.id_usuario, u.nombre, u.cedula, u.cargo,
                   a.nombre as area, e.nombre as empresa,
                   u.lap_marca, u.lap_modelo, u.lap_serie, u.lap_estado, 
                   u.sts, u.fecha_asignacion
            FROM usuarios u
            LEFT JOIN areas a ON u.id_area = a.id_area
            LEFT JOIN empresas e ON u.id_empresa = e.id_empresa
            WHERE u.nombre LIKE ?
        ''', (f'%{nombre}%',))
        
        resultado = cursor.fetchone()
        
        if not resultado:
            conn.close()
            return None
        
        datos = {
            'id_usuario': resultado[0],
            'nombre_usuario': resultado[1],
            'cedula': resultado[2] or '',
            'cargo': resultado[3] or '',
            'area': resultado[4] or '',
            'empresa': resultado[5] or 'VUBA LOGISTICS',
            'lap_marca': resultado[6] or '',
            'lap_modelo': resultado[7] or '',
            'lap_serie': resultado[8] or '',
            'lap_estado': resultado[9] or '',
            'sts': resultado[10] or 'ACTIVO',
            'fecha_asignacion': resultado[11] or '',
            'fecha': datetime.now().strftime('%d/%m/%Y'),
            'numero_responsiva': f"RSP-{datetime.now().strftime('%Y%m%d')}-{resultado[0]:04d}",
            'jefe_nombre': 'JEFE INMEDIATO',
            'ti_nombre': 'DEPARTAMENTO TI',
            'accesorios': []
        }
        
        # Obtener accesorios
        cursor.execute('''
            SELECT tipo, marca, modelo, serie, estado
            FROM accesorios
            WHERE id_usuario = ?
        ''', (datos['id_usuario'],))
        
        for acc in cursor.fetchall():
            datos['accesorios'].append({
                'tipo': acc[0],
                'marca': acc[1] or '',
                'modelo': acc[2] or '',
                'serie': acc[3] or '',
                'estado': acc[4] or 'BUENO'
            })
        
        conn.close()
        return datos


def generar_excel_simple(datos, ruta_salida):
    """Genera Excel simple para demostración"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Responsiva"
    
    # Estilos básicos
    titulo_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    titulo_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    label_font = Font(name='Calibri', size=10, bold=True)
    
    fila = 1
    
    # Título
    ws.merge_cells(f'A{fila}:F{fila}')
    cell = ws[f'A{fila}']
    cell.value = 'ACTA DE RECEPCIÓN DE EQUIPOS DE CÓMPUTO'
    cell.font = titulo_font
    cell.fill = titulo_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    fila += 2
    
    # Información
    ws[f'A{fila}'].value = 'Fecha:'
    ws[f'A{fila}'].font = label_font
    ws[f'B{fila}'].value = datos.get('fecha', '')
    fila += 1
    
    ws[f'A{fila}'].value = 'No. Responsiva:'
    ws[f'A{fila}'].font = label_font
    ws[f'B{fila}'].value = datos.get('numero_responsiva', '')
    fila += 2
    
    # Empleado
    ws[f'A{fila}'].value = 'Nombre:'
    ws[f'A{fila}'].font = label_font
    ws.merge_cells(f'B{fila}:D{fila}')
    ws[f'B{fila}'].value = datos.get('nombre_usuario', '')
    fila += 1
    
    ws[f'A{fila}'].value = 'Área:'
    ws[f'A{fila}'].font = label_font
    ws[f'B{fila}'].value = datos.get('area', '')
    fila += 2
    
    # Equipo
    ws[f'A{fila}'].value = 'Equipo:'
    ws[f'A{fila}'].font = label_font
    ws.merge_cells(f'B{fila}:D{fila}')
    ws[f'B{fila}'].value = f"{datos.get('lap_marca', '')} {datos.get('lap_modelo', '')}"
    fila += 1
    
    ws[f'A{fila}'].value = 'Serie:'
    ws[f'A{fila}'].font = label_font
    ws[f'B{fila}'].value = datos.get('lap_serie', '')
    fila += 1
    
    ws[f'A{fila}'].value = 'Estado:'
    ws[f'A{fila}'].font = label_font
    ws[f'B{fila}'].value = datos.get('lap_estado', '')
    
    wb.save(ruta_salida)
    return ruta_salida


def main():
    """Función principal de demostración"""
    print("=" * 60)
    print("DEMOSTRACIÓN - SISTEMA DE RESPONSIVAS")
    print("=" * 60)
    print()
    
    # Crear base de datos
    print("1. Creando base de datos de ejemplo...")
    db = DatabaseManager()
    print("   ✓ Base de datos creada")
    print()
    
    # Buscar usuario
    print("2. Buscando empleado 'Juan'...")
    datos = db.buscar_usuario("Juan")
    
    if datos:
        print(f"   ✓ Empleado encontrado: {datos['nombre_usuario']}")
        print(f"   ✓ Equipo: {datos['lap_marca']} {datos['lap_modelo']}")
        print(f"   ✓ Accesorios: {len(datos['accesorios'])} items")
    else:
        print("   ✗ No se encontró el empleado")
        return
    
    print()
    
    # Crear carpeta de salida
    output_dir = Path("Demo_Responsivas")
    output_dir.mkdir(exist_ok=True)
    
    # Generar Excel
    print("3. Generando documento Excel...")
    nombre_excel = f"Responsiva_Demo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta_excel = output_dir / nombre_excel
    
    try:
        generar_excel_simple(datos, str(ruta_excel))
        size_excel = ruta_excel.stat().st_size
        print(f"   ✓ Excel generado: {nombre_excel}")
        print(f"   ✓ Tamaño: {size_excel:,} bytes")
        print(f"   ✓ Ubicación: {output_dir}/")
    except Exception as e:
        print(f"   ✗ Error: {str(e)}")
    
    print()
    
    # Generar PDF (si está disponible)
    if REPORTLAB_AVAILABLE:
        print("4. Generando documento PDF...")
        print("   ⚠️  Implementación completa disponible en sistema_responsivas.py")
    else:
        print("4. ReportLab no disponible - PDF omitido")
    
    print()
    print("=" * 60)
    print("DEMOSTRACIÓN COMPLETADA")
    print("=" * 60)
    print()
    print(f"Archivos generados en: {output_dir.absolute()}/")
    print()
    print("Para usar la versión completa con interfaz gráfica:")
    print("    python sistema_responsivas.py")
    print()
    
    # Limpiar base de datos de demo
    Path('responsivas_demo.db').unlink(missing_ok=True)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n✗ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

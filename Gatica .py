import tkinter as tk
from tkinter import ttk, messagebox
import pyodbc
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
except ImportError:
    print("Instala ReportLab: pip install reportlab")


class GeneradorResponsivas:
    """Genera Excel y PDF de responsivas sin usar XML"""
    
    def __init__(self, base_dir):
        self.base_dir = base_dir
        self.output_dir = os.path.join(base_dir, "Responsivas_Generadas")
        os.makedirs(self.output_dir, exist_ok=True)
        self.server = 'ADMIN\\SAA'
        self.database = 'TI_VUBA'
        self._crear_estilos_pdf()

    def get_connection(self):
        """Conecta a SQL Server"""
        return pyodbc.connect(
            f'DRIVER={{SQL Server}};SERVER={self.server};DATABASE={self.database};Trusted_Connection=yes;')

    def _crear_estilos_pdf(self):
        """Crea estilos personalizados para PDF"""
        self.styles = getSampleStyleSheet()
        
        self.styles.add(ParagraphStyle(
            name='Titulo',
            parent=self.styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        self.styles.add(ParagraphStyle(
            name='Seccion',
            parent=self.styles['Heading2'],
            fontSize=12,
            textColor=colors.white,
            backColor=colors.HexColor('#1F4E78'),
            spaceAfter=6,
            spaceBefore=12,
            fontName='Helvetica-Bold',
            alignment=TA_LEFT,
            leftIndent=5
        ))
        
        self.styles.add(ParagraphStyle(
            name='Normal_Small',
            parent=self.styles['Normal'],
            fontSize=9,
            spaceAfter=3,
            leading=11
        ))

    def obtener_datos_usuario(self, nombre_usuario):
        """Lee datos del empleado desde SQL Server"""
        try:
            conn = self.get_connection()
            cur = conn.cursor()
            
            cur.execute("""
                SELECT U.ID_usuario, U.nombre, A.descripcion, E.descripcion, 
                       U.LAP_MARCA, U.LAP_MODELO, U.LAP_SERIE, U.LAP_ESTADO, U.STS
                FROM usuarios U 
                JOIN Area A ON U.ID_area = A.ID_area 
                JOIN Empresa E ON U.ID_empresa = E.ID_empresa
                WHERE U.nombre LIKE ?
            """, (f'%{nombre_usuario}%',))
            
            resultado = cur.fetchone()
            if not resultado:
                conn.close()
                return None
            
            id_usuario = resultado[0]
            datos = {
                'nombre_usuario': resultado[1],
                'area': resultado[2],
                'empresa': resultado[3],
                'lap_marca': resultado[4],
                'lap_modelo': resultado[5],
                'lap_serie': resultado[6],
                'lap_estado': resultado[7],
                'sts': resultado[8],
                'fecha': datetime.now().strftime('%d/%m/%Y'),
                'numero_responsiva': f"RSP-{datetime.now().strftime('%Y%m%d')}-{id_usuario}",
                'cargo': '',
                'cedula': '',
                'jefe_nombre': 'JEFE INMEDIATO',
                'ti_nombre': 'DEPARTAMENTO TI',
                'accesorios': []
            }
            
            # Obtener accesorios
            cur.execute("""
                SELECT TIPO_ACCESORIO, NO_SERIE, ESTADO_FISICO 
                FROM TI_DETALLE_ACCESORIOS 
                WHERE ID_INVENTARIO_PRINCIPAL = ?
            """, (id_usuario,))
            
            for acc in cur.fetchall():
                datos['accesorios'].append({
                    'tipo': acc[0],
                    'serie': acc[1],
                    'estado': acc[2]
                })
            
            conn.close()
            return datos
            
        except Exception as e:
            print(f"Error: {e}")
            return None

    def generar_excel(self, datos, nombre_archivo=None):
        """Genera Excel profesional"""
        if nombre_archivo is None:
            nombre_archivo = f"Responsiva_{datos.get('nombre_usuario', 'Empleado').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        ruta_salida = os.path.join(self.output_dir, nombre_archivo)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Responsiva"
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        
        # Estilos
        titulo_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        titulo_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        encabezado_font = Font(name='Calibri', size=11, bold=True)
        encabezado_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        normal_font = Font(name='Calibri', size=10)
        
        # Ancho columnas
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        fila = 1
        
        # ENCABEZADO
        ws.merge_cells(f'A{fila}:H{fila}')
        cell = ws[f'A{fila}']
        cell.value = 'ACTA DE RECEPCIÓN Y RESPONSABILIDAD DE EQUIPOS DE CÓMPUTO'
        cell.font = titulo_font
        cell.fill = titulo_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[fila].height = 25
        fila += 1
        
        ws.merge_cells(f'A{fila}:H{fila}')
        cell = ws[f'A{fila}']
        cell.value = f"Empresa: {datos.get('empresa', 'VUBA LOGISTICS')}"
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center')
        fila += 2
        
        # INFORMACIÓN GENERAL
        ws.merge_cells(f'A{fila}:H{fila}')
        cell = ws[f'A{fila}']
        cell.value = 'INFORMACIÓN GENERAL'
        cell.font = encabezado_font
        cell.fill = encabezado_fill
        fila += 1
        
        ws[f'A{fila}'].value = 'Fecha:'
        ws[f'B{fila}'].value = datos.get('fecha', '')
        ws[f'E{fila}'].value = 'No. Responsiva:'
        ws[f'F{fila}'].value = datos.get('numero_responsiva', '')
        fila += 2
        
        # DATOS DEL EMPLEADO
        ws.merge_cells(f'A{fila}:H{fila}')
        cell = ws[f'A{fila}']
        cell.value = 'DATOS DEL EMPLEADO'
        cell.font = encabezado_font
        cell.fill = encabezado_fill
        fila += 1
        
        ws[f'A{fila}'].value = 'Nombre:'
        ws.merge_cells(f'B{fila}:D{fila}')
        ws[f'B{fila}'].value = datos.get('nombre_usuario', '')
        ws[f'E{fila}'].value = 'Cédula:'
        ws.merge_cells(f'F{fila}:H{fila}')
        ws[f'F{fila}'].value = datos.get('cedula', '')
        fila += 1
        
        ws[f'A{fila}'].value = 'Cargo:'
        ws.merge_cells(f'B{fila}:D{fila}')
        ws[f'B{fila}'].value = datos.get('cargo', '')
        ws[f'E{fila}'].value = 'Área:'
        ws.merge_cells(f'F{fila}:H{fila}')
        ws[f'F{fila}'].value = datos.get('area', '')
        fila += 2
        
        # EQUIPO PRINCIPAL
        ws.merge_cells(f'A{fila}:H{fila}')
        cell = ws[f'A{fila}']
        cell.value = 'EQUIPO PRINCIPAL ASIGNADO'
        cell.font = encabezado_font
        cell.fill = encabezado_fill
        fila += 1
        
        ws[f'A{fila}'].value = 'Marca:'
        ws[f'B{fila}'].value = datos.get('lap_marca', '')
        ws[f'C{fila}'].value = 'Modelo:'
        ws[f'D{fila}'].value = datos.get('lap_modelo', '')
        ws[f'E{fila}'].value = 'No. Serie:'
        ws.merge_cells(f'F{fila}:H{fila}')
        ws[f'F{fila}'].value = datos.get('lap_serie', '')
        fila += 1
        
        ws[f'A{fila}'].value = 'Estado Físico:'
        ws[f'B{fila}'].value = datos.get('lap_estado', '')
        ws[f'C{fila}'].value = 'Estatus:'
        ws[f'D{fila}'].value = datos.get('sts', '')
        fila += 2
        
        # ACCESORIOS
        if datos.get('accesorios'):
            ws.merge_cells(f'A{fila}:H{fila}')
            cell = ws[f'A{fila}']
            cell.value = 'ACCESORIOS ASIGNADOS'
            cell.font = encabezado_font
            cell.fill = encabezado_fill
            fila += 1
            
            for acc in datos['accesorios']:
                ws[f'A{fila}'].value = acc.get('tipo', '')
                ws[f'C{fila}'].value = 'Serie:'
                ws.merge_cells(f'D{fila}:H{fila}')
                ws[f'D{fila}'].value = acc.get('serie', '')
                fila += 1
            fila += 1
        
        # RESPONSABILIDADES
        ws.merge_cells(f'A{fila}:H{fila}')
        cell = ws[f'A{fila}']
        cell.value = 'RESPONSABILIDADES DEL EMPLEADO'
        cell.font = encabezado_font
        cell.fill = encabezado_fill
        fila += 1
        
        responsabilidades = [
            '• El empleado es responsable por el cuidado, conservación y mantenimiento del equipo.',
            '• El equipo debe ser devuelto en las mismas condiciones de recepción.',
            '• Cualquier daño por negligencia será responsabilidad del empleado.',
            '• El equipo es propiedad de la empresa y debe ser usado únicamente para fines laborales.',
            '• Reportar inmediatamente cualquier daño o mal funcionamiento.'
        ]
        
        for resp in responsabilidades:
            ws.merge_cells(f'A{fila}:H{fila}')
            cell = ws[f'A{fila}']
            cell.value = resp
            cell.font = Font(name='Calibri', size=9)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[fila].height = 20
            fila += 1
        
        fila += 2
        
        # FIRMAS
        ws.merge_cells(f'A{fila}:H{fila}')
        cell = ws[f'A{fila}']
        cell.value = 'FIRMAS DE RESPONSABILIDAD'
        cell.font = encabezado_font
        cell.fill = encabezado_fill
        fila += 2
        
        fecha = datos.get('fecha', '')
        nombre = datos.get('nombre_usuario', '')
        jefe = datos.get('jefe_nombre', '')
        ti = datos.get('ti_nombre', '')
        
        ws.merge_cells(f'A{fila}:C{fila}')
        ws[f'A{fila}'].value = f"________________________\n{nombre}\nFecha: {fecha}"
        ws[f'A{fila}'].alignment = Alignment(horizontal='center', wrap_text=True, vertical='top')
        
        ws.merge_cells(f'D{fila}:F{fila}')
        ws[f'D{fila}'].value = f"________________________\n{jefe}\nFecha: {fecha}"
        ws[f'D{fila}'].alignment = Alignment(horizontal='center', wrap_text=True, vertical='top')
        
        ws.merge_cells(f'G{fila}:H{fila}')
        ws[f'G{fila}'].value = f"________________________\n{ti}\nFecha: {fecha}"
        ws[f'G{fila}'].alignment = Alignment(horizontal='center', wrap_text=True, vertical='top')
        
        ws.row_dimensions[fila].height = 60
        
        wb.save(ruta_salida)
        return ruta_salida

    def generar_pdf(self, datos, nombre_archivo=None):
        """Genera PDF profesional"""
        if nombre_archivo is None:
            nombre_archivo = f"Responsiva_{datos.get('nombre_usuario', 'Empleado').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        ruta_salida = os.path.join(self.output_dir, nombre_archivo)
        
        doc = SimpleDocTemplate(
            ruta_salida,
            pagesize=letter,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        story = []
        
        # ENCABEZADO
        story.append(Paragraph('VUBA LOGISTICS', self.styles['Titulo']))
        story.append(Paragraph('ACTA DE RECEPCIÓN Y RESPONSABILIDAD DE EQUIPOS DE CÓMPUTO', self.styles['Titulo']))
        story.append(Spacer(1, 0.2*inch))
        
        # INFORMACIÓN GENERAL
        fecha = datos.get('fecha', '')
        numero = datos.get('numero_responsiva', '')
        
        tabla_info = Table([['Fecha:', fecha, 'No. Responsiva:', numero]], colWidths=[1.2*inch, 2*inch, 1.8*inch, 2*inch])
        tabla_info.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(tabla_info)
        story.append(Spacer(1, 0.15*inch))
        
        # DATOS EMPLEADO
        story.append(Paragraph('DATOS DEL EMPLEADO', self.styles['Seccion']))
        
        nombre = datos.get('nombre_usuario', '')
        cedula = datos.get('cedula', '')
        cargo = datos.get('cargo', '')
        area = datos.get('area', '')
        
        tabla_emp = Table([
            ['Nombre:', nombre, 'Cédula:', cedula],
            ['Cargo:', cargo, 'Área:', area]
        ], colWidths=[1.2*inch, 2*inch, 1.2*inch, 2*inch])
        tabla_emp.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(tabla_emp)
        story.append(Spacer(1, 0.15*inch))
        
        # EQUIPO
        story.append(Paragraph('EQUIPO PRINCIPAL ASIGNADO', self.styles['Seccion']))
        
        marca = datos.get('lap_marca', '')
        modelo = datos.get('lap_modelo', '')
        serie = datos.get('lap_serie', '')
        estado = datos.get('lap_estado', '')
        
        tabla_eq = Table([
            ['Marca:', marca, 'Modelo:', modelo],
            ['No. Serie:', serie, 'Estado Físico:', estado]
        ], colWidths=[1.2*inch, 2*inch, 1.2*inch, 2*inch])
        tabla_eq.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(tabla_eq)
        story.append(Spacer(1, 0.15*inch))
        
        # ACCESORIOS
        if datos.get('accesorios'):
            story.append(Paragraph('ACCESORIOS ASIGNADOS', self.styles['Seccion']))
            
            acc_data = [['Tipo', 'No. Serie', 'Estado']]
            for acc in datos['accesorios']:
                acc_data.append([acc.get('tipo', ''), acc.get('serie', ''), acc.get('estado', '')])
            
            tabla_acc = Table(acc_data, colWidths=[2*inch, 2.5*inch, 1.5*inch])
            tabla_acc.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(tabla_acc)
            story.append(Spacer(1, 0.15*inch))
        
        # RESPONSABILIDADES
        story.append(Paragraph('RESPONSABILIDADES DEL EMPLEADO', self.styles['Seccion']))
        
        for texto in [
            '• El empleado es responsable por el cuidado y conservación del equipo.',
            '• El equipo debe ser devuelto en las mismas condiciones de recepción.',
            '• Cualquier daño por negligencia será responsabilidad del empleado.',
            '• El equipo es propiedad de la empresa y debe ser usado únicamente para fines laborales.',
            '• Reportar inmediatamente cualquier daño o mal funcionamiento.'
        ]:
            story.append(Paragraph(texto, self.styles['Normal_Small']))
        
        story.append(Spacer(1, 0.2*inch))
        
        # FIRMAS
        story.append(Paragraph('FIRMAS DE RESPONSABILIDAD', self.styles['Seccion']))
        story.append(Spacer(1, 0.15*inch))
        
        tabla_firmas = Table([
            [
                f"_____________________\n{nombre}\nFecha: {fecha}",
                f"_____________________\n{datos.get('jefe_nombre', 'JEFE')}\nFecha: {fecha}",
                f"_____________________\n{datos.get('ti_nombre', 'TI')}\nFecha: {fecha}"
            ]
        ], colWidths=[2*inch, 2*inch, 2*inch])
        tabla_firmas.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 25),
        ]))
        story.append(tabla_firmas)
        
        # NOTAS
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph('• Este documento debe ser guardado como comprobante.', self.styles['Normal_Small']))
        story.append(Paragraph('• Se requieren las tres firmas para validez.', self.styles['Normal_Small']))
        
        try:
            doc.build(story)
            return ruta_salida
        except Exception as e:
            print(f"Error PDF: {e}")
            return None


class InventarioApp:
    """Interfaz gráfica para generar responsivas"""
    
    def __init__(self, master):
        self.master = master
        master.title("GENERADOR DE RESPONSIVAS - VUBA LOGISTICS")
        master.geometry("600x500")
        master.resizable(False, False)
        
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.generador = GeneradorResponsivas(self.base_dir)
        self.datos_actual = None
        
        self.setup_ui()

    def setup_ui(self):
        # ENCABEZADO
        header = ttk.Frame(self.master, padding=15)
        header.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(header, text="GENERADOR DE RESPONSIVAS", font=('Calibri', 18, 'bold')).pack()
        ttk.Label(header, text="Departamento TI - VUBA LOGISTICS", font=('Calibri', 10)).pack()
        
        # FRAME PRINCIPAL
        main_frame = ttk.Frame(self.master, padding=20)
        main_frame.pack(fill="both", expand=True)
        
        # BÚSQUEDA
        ttk.Label(main_frame, text="Nombre del Empleado:", font=('Calibri', 11, 'bold')).pack(anchor='w', pady=(0, 5))
        
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill="x", pady=(0, 15))
        
        self.ent_nombre = ttk.Entry(search_frame, width=40, font=('Calibri', 11))
        self.ent_nombre.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.ent_nombre.bind('<Return>', lambda e: self.buscar_empleado())
        
        ttk.Button(search_frame, text="🔍 Buscar", command=self.buscar_empleado).pack(side="left")
        
        ttk.Separator(main_frame, orient='horizontal').pack(fill="x", pady=15)
        
        # INFORMACIÓN
        ttk.Label(main_frame, text="Datos del Empleado:", font=('Calibri', 11, 'bold')).pack(anchor='w', pady=(0, 10))
        
        self.info_frame = ttk.LabelFrame(main_frame, text="", padding=10)
        self.info_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        self.lbl_info = ttk.Label(self.info_frame, text="Busca un empleado para ver sus datos", font=('Calibri', 10), justify='left')
        self.lbl_info.pack(anchor='nw')
        
        # BOTONES
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Button(button_frame, text="📄 Generar Excel", command=self.generar_excel, width=20).pack(side="left", padx=5)
        ttk.Button(button_frame, text="📋 Generar PDF", command=self.generar_pdf, width=20).pack(side="left", padx=5)
        
        # STATUS
        self.lbl_status = ttk.Label(self.master, text="Listo", relief="sunken", font=('Calibri', 9))
        self.lbl_status.pack(fill="x", side="bottom")

    def buscar_empleado(self):
        nombre = self.ent_nombre.get().strip()
        if not nombre:
            messagebox.showwarning("Atención", "Ingresa el nombre del empleado")
            return
        
        try:
            self.datos_actual = self.generador.obtener_datos_usuario(nombre)
            
            if not self.datos_actual:
                messagebox.showerror("Error", f"No encontrado: '{nombre}'")
                self.lbl_info.config(text="Empleado no encontrado")
                self.datos_actual = None
                return
            
            info_text = f"""Nombre: {self.datos_actual.get('nombre_usuario', 'N/A')}
Área: {self.datos_actual.get('area', 'N/A')}
Empresa: {self.datos_actual.get('empresa', 'N/A')}

Equipo:
  Marca: {self.datos_actual.get('lap_marca', 'N/A')}
  Modelo: {self.datos_actual.get('lap_modelo', 'N/A')}
  Serial: {self.datos_actual.get('lap_serie', 'N/A')}
  Estado: {self.datos_actual.get('lap_estado', 'N/A')}

Accesorios: {len(self.datos_actual.get('accesorios', []))} artículos"""
            
            self.lbl_info.config(text=info_text)
            self.lbl_status.config(text="✓ Empleado encontrado. Listo para generar.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error: {str(e)}")
            self.lbl_status.config(text="❌ Error")

    def generar_excel(self):
        if not self.datos_actual:
            messagebox.showwarning("Atención", "Primero busca un empleado")
            return
        
        try:
            self.lbl_status.config(text="Generando Excel...")
            self.master.update()
            
            ruta = self.generador.generar_excel(self.datos_actual)
            self.lbl_status.config(text="✓ Excel generado")
            
            messagebox.showinfo("Éxito", f"✓ Excel generado:\n\n{os.path.basename(ruta)}\n\nCarpeta: Responsivas_Generadas/")
        except Exception as e:
            messagebox.showerror("Error", f"Error: {str(e)}")

    def generar_pdf(self):
        if not self.datos_actual:
            messagebox.showwarning("Atención", "Primero busca un empleado")
            return
        
        try:
            self.lbl_status.config(text="Generando PDF...")
            self.master.update()
            
            ruta = self.generador.generar_pdf(self.datos_actual)
            self.lbl_status.config(text="✓ PDF generado")
            
            messagebox.showinfo("Éxito", f"✓ PDF generado:\n\n{os.path.basename(ruta)}\n\nCarpeta: Responsivas_Generadas/")
        except Exception as e:
            messagebox.showerror("Error", f"Error: {str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = InventarioApp(root)
    root.mainloop()

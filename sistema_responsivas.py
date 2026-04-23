"""
Sistema de Gestión de Responsivas - VUBA LOGISTICS
Base de datos SQLite local con exportación a Excel y PDF

Autor: Departamento TI - VUBA LOGISTICS
Versión: 2.0
Fecha: Enero 2025

Descripción:
    Sistema para la gestión de responsivas de equipos de cómputo.
    Permite registrar empleados, equipos, accesorios y generar documentos
    de responsabilidad en formato Excel y PDF.

Módulos principales:
    - DatabaseManager: Gestión de base de datos SQLite
    - ExcelGenerator: Generación de archivos Excel
    - PDFGenerator: Generación de archivos PDF
    - AplicacionResponsivas: Interfaz gráfica principal

Dependencias:
    - tkinter: Interfaz gráfica (incluido en Python)
    - sqlite3: Base de datos (incluido en Python)
    - openpyxl: Generación de Excel
    - reportlab: Generación de PDF
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import os
from datetime import datetime
from pathlib import Path

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("ADVERTENCIA: ReportLab no instalado. Ejecuta: pip install reportlab")


class DatabaseManager:
    """Gestiona la base de datos SQLite local"""
    
    def __init__(self, db_path='responsivas.db'):
        self.db_path = db_path
        self.init_database()
    
    def get_connection(self):
        """Crea conexión a la base de datos"""
        return sqlite3.connect(self.db_path)
    
    def init_database(self):
        """Inicializa las tablas si no existen"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Tabla de empresas
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS empresas (
                id_empresa INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL UNIQUE,
                fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Tabla de áreas
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS areas (
                id_area INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL UNIQUE,
                descripcion TEXT
            )
        ''')
        
        # Tabla de usuarios/empleados
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS usuarios (
                id_usuario INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                cedula TEXT,
                cargo TEXT,
                id_area INTEGER,
                id_empresa INTEGER,
                lap_marca TEXT,
                lap_modelo TEXT,
                lap_serie TEXT UNIQUE,
                lap_estado TEXT,
                sts TEXT DEFAULT 'ACTIVO',
                fecha_asignacion DATE,
                FOREIGN KEY (id_area) REFERENCES areas (id_area),
                FOREIGN KEY (id_empresa) REFERENCES empresas (id_empresa)
            )
        ''')
        
        # Tabla de accesorios
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS accesorios (
                id_accesorio INTEGER PRIMARY KEY AUTOINCREMENT,
                id_usuario INTEGER,
                tipo TEXT NOT NULL,
                marca TEXT,
                modelo TEXT,
                serie TEXT,
                estado TEXT DEFAULT 'BUENO',
                fecha_asignacion DATE,
                FOREIGN KEY (id_usuario) REFERENCES usuarios (id_usuario)
            )
        ''')
        
        # Tabla de responsivas generadas
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS responsivas (
                id_responsiva INTEGER PRIMARY KEY AUTOINCREMENT,
                numero_responsiva TEXT UNIQUE,
                id_usuario INTEGER,
                fecha_generacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                tipo_documento TEXT,
                ruta_archivo TEXT,
                FOREIGN KEY (id_usuario) REFERENCES usuarios (id_usuario)
            )
        ''')
        
        conn.commit()
        conn.close()
        
        # Insertar datos de ejemplo si está vacío
        self._insert_sample_data()
    
    def _insert_sample_data(self):
        """Inserta datos de ejemplo si la base está vacía"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM empresas")
        if cursor.fetchone()[0] == 0:
            # Empresas
            cursor.execute("INSERT INTO empresas (nombre) VALUES (?)", ("VUBA LOGISTICS",))
            
            # Áreas
            areas = ["Tecnología", "Administración", "Recursos Humanos", "Operaciones", "Logística"]
            for area in areas:
                cursor.execute("INSERT INTO areas (nombre) VALUES (?)", (area,))
            
            # Usuarios de ejemplo
            usuarios = [
                ("Juan Pérez García", "12345678", "Analista TI", 1, 1, "DELL", "Latitude 5420", "SN-DELL-2024-001", "EXCELENTE", "2024-01-15"),
                ("María López Martínez", "87654321", "Gerente Admin", 2, 1, "HP", "EliteBook 840", "SN-HP-2024-002", "BUENO", "2024-02-10"),
                ("Carlos Ramírez Torres", "11223344", "Jefe RRHH", 3, 1, "LENOVO", "ThinkPad X1", "SN-LENOVO-2024-003", "EXCELENTE", "2024-03-05"),
            ]
            
            for usuario in usuarios:
                cursor.execute('''
                    INSERT INTO usuarios (nombre, cedula, cargo, id_area, id_empresa, 
                                        lap_marca, lap_modelo, lap_serie, lap_estado, fecha_asignacion)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', usuario)
            
            # Accesorios de ejemplo
            accesorios = [
                (1, "MOUSE", "Logitech", "M720", "SER-MOUSE-001", "BUENO", "2024-01-15"),
                (1, "TECLADO", "Logitech", "K380", "SER-TECLADO-001", "BUENO", "2024-01-15"),
                (2, "MOUSE", "HP", "Wireless", "SER-MOUSE-002", "EXCELENTE", "2024-02-10"),
                (3, "MOUSE", "Lenovo", "M300", "SER-MOUSE-003", "BUENO", "2024-03-05"),
            ]
            
            for acc in accesorios:
                cursor.execute('''
                    INSERT INTO accesorios (id_usuario, tipo, marca, modelo, serie, estado, fecha_asignacion)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', acc)
            
            conn.commit()
        
        conn.close()
    
    def buscar_usuario(self, nombre):
        """Busca un usuario por nombre"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT u.id_usuario, u.nombre, u.cedula, u.cargo,
                   a.nombre as area, e.nombre as empresa,
                   u.lap_marca, u.lap_modelo, u.lap_serie, u.lap_estado, 
                   u.sts, u.fecha_asignacion
            FROM usuarios u
            LEFT JOIN areas a ON u.id_area = a.id_area
            LEFT JOIN empresas e ON u.id_empresa = e.id_empresa
            WHERE u.nombre LIKE ?
        ''', (f'%{nombre}%',))
        
        resultado = cursor.fetchone()
        
        if not resultado:
            conn.close()
            return None
        
        datos = {
            'id_usuario': resultado[0],
            'nombre_usuario': resultado[1],
            'cedula': resultado[2] or '',
            'cargo': resultado[3] or '',
            'area': resultado[4] or '',
            'empresa': resultado[5] or 'VUBA LOGISTICS',
            'lap_marca': resultado[6] or '',
            'lap_modelo': resultado[7] or '',
            'lap_serie': resultado[8] or '',
            'lap_estado': resultado[9] or '',
            'sts': resultado[10] or 'ACTIVO',
            'fecha_asignacion': resultado[11] or '',
            'fecha': datetime.now().strftime('%d/%m/%Y'),
            'numero_responsiva': f"RSP-{datetime.now().strftime('%Y%m%d')}-{resultado[0]:04d}",
            'jefe_nombre': 'JEFE INMEDIATO',
            'ti_nombre': 'DEPARTAMENTO TI',
            'accesorios': []
        }
        
        # Obtener accesorios
        cursor.execute('''
            SELECT tipo, marca, modelo, serie, estado
            FROM accesorios
            WHERE id_usuario = ?
        ''', (datos['id_usuario'],))
        
        for acc in cursor.fetchall():
            datos['accesorios'].append({
                'tipo': acc[0],
                'marca': acc[1] or '',
                'modelo': acc[2] or '',
                'serie': acc[3] or '',
                'estado': acc[4] or 'BUENO'
            })
        
        conn.close()
        return datos
    
    def listar_usuarios(self):
        """Lista todos los usuarios"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT u.id_usuario, u.nombre, u.cedula, u.cargo, a.nombre as area
            FROM usuarios u
            LEFT JOIN areas a ON u.id_area = a.id_area
            ORDER BY u.nombre
        ''')
        
        usuarios = cursor.fetchall()
        conn.close()
        return usuarios
    
    def registrar_responsiva(self, numero_responsiva, id_usuario, tipo_documento, ruta_archivo):
        """
        Registra una responsiva generada en el historial
        
        Args:
            numero_responsiva (str): Número único de la responsiva
            id_usuario (int): ID del usuario asociado
            tipo_documento (str): Tipo de documento (EXCEL o PDF)
            ruta_archivo (str): Ruta completa del archivo generado
            
        Returns:
            bool: True si se registró correctamente, False si hubo error
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO responsivas (numero_responsiva, id_usuario, tipo_documento, ruta_archivo)
                VALUES (?, ?, ?, ?)
            ''', (numero_responsiva, id_usuario, tipo_documento, ruta_archivo))
            conn.commit()
            conn.close()
            return True
        except sqlite3.IntegrityError:
            conn.close()
            return False
    
    def agregar_empleado(self, nombre, cedula, cargo, id_area, id_empresa, 
                        lap_marca, lap_modelo, lap_serie, lap_estado):
        """
        Agrega un nuevo empleado a la base de datos
        
        Args:
            nombre (str): Nombre completo del empleado
            cedula (str): Número de cédula
            cargo (str): Cargo del empleado
            id_area (int): ID del área
            id_empresa (int): ID de la empresa
            lap_marca (str): Marca del equipo
            lap_modelo (str): Modelo del equipo
            lap_serie (str): Número de serie del equipo
            lap_estado (str): Estado físico del equipo
            
        Returns:
            tuple: (bool, str) - (éxito, mensaje)
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            fecha_actual = datetime.now().strftime('%Y-%m-%d')
            cursor.execute('''
                INSERT INTO usuarios (nombre, cedula, cargo, id_area, id_empresa,
                                    lap_marca, lap_modelo, lap_serie, lap_estado, 
                                    sts, fecha_asignacion)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'ACTIVO', ?)
            ''', (nombre, cedula, cargo, id_area, id_empresa, 
                  lap_marca, lap_modelo, lap_serie, lap_estado, fecha_actual))
            
            id_usuario = cursor.lastrowid
            conn.commit()
            conn.close()
            return True, f"Empleado agregado correctamente con ID: {id_usuario}"
        except sqlite3.IntegrityError as e:
            conn.close()
            if 'lap_serie' in str(e):
                return False, "Error: El número de serie ya existe en el sistema"
            return False, f"Error al agregar empleado: {str(e)}"
        except Exception as e:
            conn.close()
            return False, f"Error inesperado: {str(e)}"
    
    def agregar_accesorio(self, id_usuario, tipo, marca, modelo, serie, estado):
        """
        Agrega un accesorio a un empleado
        
        Args:
            id_usuario (int): ID del usuario
            tipo (str): Tipo de accesorio (MOUSE, TECLADO, etc.)
            marca (str): Marca del accesorio
            modelo (str): Modelo del accesorio
            serie (str): Número de serie
            estado (str): Estado físico
            
        Returns:
            tuple: (bool, str) - (éxito, mensaje)
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            fecha_actual = datetime.now().strftime('%Y-%m-%d')
            cursor.execute('''
                INSERT INTO accesorios (id_usuario, tipo, marca, modelo, serie, estado, fecha_asignacion)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (id_usuario, tipo, marca, modelo, serie, estado, fecha_actual))
            conn.commit()
            conn.close()
            return True, "Accesorio agregado correctamente"
        except Exception as e:
            conn.close()
            return False, f"Error al agregar accesorio: {str(e)}"
    
    def listar_areas(self):
        """
        Lista todas las áreas registradas
        
        Returns:
            list: Lista de tuplas (id_area, nombre)
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id_area, nombre FROM areas ORDER BY nombre")
        areas = cursor.fetchall()
        conn.close()
        return areas
    
    def listar_empresas(self):
        """
        Lista todas las empresas registradas
        
        Returns:
            list: Lista de tuplas (id_empresa, nombre)
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id_empresa, nombre FROM empresas ORDER BY nombre")
        empresas = cursor.fetchall()
        conn.close()
        return empresas
    
    def agregar_area(self, nombre, descripcion=""):
        """
        Agrega una nueva área
        
        Args:
            nombre (str): Nombre del área
            descripcion (str): Descripción opcional
            
        Returns:
            tuple: (bool, str) - (éxito, mensaje)
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO areas (nombre, descripcion)
                VALUES (?, ?)
            ''', (nombre, descripcion))
            conn.commit()
            conn.close()
            return True, "Area agregada correctamente"
        except sqlite3.IntegrityError:
            conn.close()
            return False, "Error: El área ya existe"
        except Exception as e:
            conn.close()
            return False, f"Error al agregar área: {str(e)}"
    
    def agregar_empresa(self, nombre):
        """
        Agrega una nueva empresa
        
        Args:
            nombre (str): Nombre de la empresa
            
        Returns:
            tuple: (bool, str) - (éxito, mensaje)
        """
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO empresas (nombre)
                VALUES (?)
            ''', (nombre,))
            conn.commit()
            conn.close()
            return True, "Empresa agregada correctamente"
        except sqlite3.IntegrityError:
            conn.close()
            return False, "Error: La empresa ya existe"
        except Exception as e:
            conn.close()
            return False, f"Error al agregar empresa: {str(e)}"


class ExcelGenerator:
    """Genera documentos Excel profesionales"""
    
    @staticmethod
    def generar(datos, ruta_salida):
        """Genera archivo Excel limpio y profesional"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Responsiva"
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        
        # Colores corporativos
        COLOR_HEADER = '1F4E78'
        COLOR_SECTION = 'D9E1F2'
        
        # Estilos
        titulo_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        titulo_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        seccion_font = Font(name='Calibri', size=11, bold=True)
        seccion_fill = PatternFill(start_color=COLOR_SECTION, end_color=COLOR_SECTION, fill_type='solid')
        normal_font = Font(name='Calibri', size=10)
        label_font = Font(name='Calibri', size=10, bold=True)
        
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Ancho de columnas
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 22
        
        fila = 1
        
        # === ENCABEZADO ===
        ws.merge_cells(f'A{fila}:F{fila}')
        cell = ws[f'A{fila}']
        cell.value = 'ACTA DE RECEPCIÓN Y RESPONSABILIDAD DE EQUIPOS DE CÓMPUTO'
        cell.font = titulo_font
        cell.fill = titulo_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[fila].height = 30
        fila += 1
        
        ws.merge_cells(f'A{fila}:F{fila}')
        cell = ws[f'A{fila}']
        cell.value = datos.get('empresa', 'VUBA LOGISTICS')
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center')
        fila += 2
        
        # === INFORMACIÓN GENERAL ===
        ws.merge_cells(f'A{fila}:F{fila}')
        cell = ws[f'A{fila}']
        cell.value = 'INFORMACIÓN GENERAL'
        cell.font = seccion_font
        cell.fill = seccion_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center')
        fila += 1
        
        ws[f'A{fila}'].value = 'Fecha:'
        ws[f'A{fila}'].font = label_font
        ws[f'B{fila}'].value = datos.get('fecha', '')
        
        ws[f'D{fila}'].value = 'No. Responsiva:'
        ws[f'D{fila}'].font = label_font
        ws.merge_cells(f'E{fila}:F{fila}')
        ws[f'E{fila}'].value = datos.get('numero_responsiva', '')
        fila += 2
        
        # === DATOS DEL EMPLEADO ===
        ws.merge_cells(f'A{fila}:F{fila}')
        cell = ws[f'A{fila}']
        cell.value = 'DATOS DEL EMPLEADO'
        cell.font = seccion_font
        cell.fill = seccion_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center')
        fila += 1
        
        ws[f'A{fila}'].value = 'Nombre Completo:'
        ws[f'A{fila}'].font = label_font
        ws.merge_cells(f'B{fila}:C{fila}')
        ws[f'B{fila}'].value = datos.get('nombre_usuario', '')
        
        ws[f'D{fila}'].value = 'Cédula:'
        ws[f'D{fila}'].font = label_font
        ws.merge_cells(f'E{fila}:F{fila}')
        ws[f'E{fila}'].value = datos.get('cedula', '')
        fila += 1
        
        ws[f'A{fila}'].value = 'Cargo:'
        ws[f'A{fila}'].font = label_font
        ws.merge_cells(f'B{fila}:C{fila}')
        ws[f'B{fila}'].value = datos.get('cargo', '')
        
        ws[f'D{fila}'].value = 'Área:'
        ws[f'D{fila}'].font = label_font
        ws.merge_cells(f'E{fila}:F{fila}')
        ws[f'E{fila}'].value = datos.get('area', '')
        fila += 2
        
        # === EQUIPO PRINCIPAL ===
        ws.merge_cells(f'A{fila}:F{fila}')
        cell = ws[f'A{fila}']
        cell.value = 'EQUIPO PRINCIPAL ASIGNADO'
        cell.font = seccion_font
        cell.fill = seccion_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center')
        fila += 1
        
        ws[f'A{fila}'].value = 'Marca:'
        ws[f'A{fila}'].font = label_font
        ws[f'B{fila}'].value = datos.get('lap_marca', '')
        
        ws[f'C{fila}'].value = 'Modelo:'
        ws[f'C{fila}'].font = label_font
        ws[f'D{fila}'].value = datos.get('lap_modelo', '')
        fila += 1
        
        ws[f'A{fila}'].value = 'Número de Serie:'
        ws[f'A{fila}'].font = label_font
        ws.merge_cells(f'B{fila}:D{fila}')
        ws[f'B{fila}'].value = datos.get('lap_serie', '')
        fila += 1
        
        ws[f'A{fila}'].value = 'Estado Físico:'
        ws[f'A{fila}'].font = label_font
        ws[f'B{fila}'].value = datos.get('lap_estado', '')
        
        ws[f'C{fila}'].value = 'Estatus:'
        ws[f'C{fila}'].font = label_font
        ws[f'D{fila}'].value = datos.get('sts', '')
        fila += 2
        
        # === ACCESORIOS ===
        if datos.get('accesorios'):
            ws.merge_cells(f'A{fila}:F{fila}')
            cell = ws[f'A{fila}']
            cell.value = 'ACCESORIOS ASIGNADOS'
            cell.font = seccion_font
            cell.fill = seccion_fill
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center')
            fila += 1
            
            # Encabezados tabla
            headers = ['Tipo', 'Marca', 'Modelo', 'Serie', 'Estado']
            cols = ['A', 'B', 'C', 'D', 'E']
            
            for col, header in zip(cols, headers):
                cell = ws[f'{col}{fila}']
                cell.value = header
                cell.font = label_font
                cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                cell.border = border_thin
                cell.alignment = Alignment(horizontal='center')
            fila += 1
            
            # Datos accesorios
            for acc in datos['accesorios']:
                ws[f'A{fila}'].value = acc.get('tipo', '')
                ws[f'B{fila}'].value = acc.get('marca', '')
                ws[f'C{fila}'].value = acc.get('modelo', '')
                ws[f'D{fila}'].value = acc.get('serie', '')
                ws[f'E{fila}'].value = acc.get('estado', '')
                
                for col in cols:
                    ws[f'{col}{fila}'].border = border_thin
                    ws[f'{col}{fila}'].alignment = Alignment(horizontal='center')
                
                fila += 1
            fila += 1
        
        # === RESPONSABILIDADES ===
        ws.merge_cells(f'A{fila}:F{fila}')
        cell = ws[f'A{fila}']
        cell.value = 'RESPONSABILIDADES DEL EMPLEADO'
        cell.font = seccion_font
        cell.fill = seccion_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center')
        fila += 1
        
        responsabilidades = [
            'El empleado es responsable del cuidado, conservación y buen uso del equipo asignado.',
            'El equipo debe ser devuelto en las mismas condiciones de recepción al finalizar la relación laboral.',
            'Cualquier daño o pérdida por negligencia será responsabilidad del empleado.',
            'El equipo es propiedad de la empresa y debe usarse únicamente para fines laborales.',
            'Debe reportar inmediatamente cualquier daño, mal funcionamiento o pérdida del equipo.'
        ]
        
        for i, resp in enumerate(responsabilidades, 1):
            ws.merge_cells(f'A{fila}:F{fila}')
            cell = ws[f'A{fila}']
            cell.value = f'{i}. {resp}'
            cell.font = Font(name='Calibri', size=9)
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='justify')
            ws.row_dimensions[fila].height = 25
            fila += 1
        
        fila += 2
        
        # === FIRMAS ===
        ws.merge_cells(f'A{fila}:F{fila}')
        cell = ws[f'A{fila}']
        cell.value = 'FIRMAS DE RESPONSABILIDAD'
        cell.font = seccion_font
        cell.fill = seccion_fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center')
        fila += 2
        
        # Espacios para firmas
        firma_empleado = f"_________________________\n{datos.get('nombre_usuario', '')}\nEMPLEADO\nFecha: {datos.get('fecha', '')}"
        firma_jefe = f"_________________________\n{datos.get('jefe_nombre', '')}\nJEFE INMEDIATO\nFecha: {datos.get('fecha', '')}"
        firma_ti = f"_________________________\n{datos.get('ti_nombre', '')}\nDEPTO. TI\nFecha: {datos.get('fecha', '')}"
        
        ws.merge_cells(f'A{fila}:B{fila}')
        ws[f'A{fila}'].value = firma_empleado
        ws[f'A{fila}'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        
        ws.merge_cells(f'C{fila}:D{fila}')
        ws[f'C{fila}'].value = firma_jefe
        ws[f'C{fila}'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        
        ws.merge_cells(f'E{fila}:F{fila}')
        ws[f'E{fila}'].value = firma_ti
        ws[f'E{fila}'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        
        ws.row_dimensions[fila].height = 80
        
        # Guardar
        wb.save(ruta_salida)
        return ruta_salida


class PDFGenerator:
    """Genera documentos PDF profesionales"""
    
    def __init__(self):
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab no está instalado")
        self._crear_estilos()
    
    def _crear_estilos(self):
        """Crea estilos para PDF"""
        self.styles = getSampleStyleSheet()
        
        self.styles.add(ParagraphStyle(
            name='TituloEmpresa',
            parent=self.styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=4,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        self.styles.add(ParagraphStyle(
            name='TituloDocumento',
            parent=self.styles['Heading1'],
            fontSize=12,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        self.styles.add(ParagraphStyle(
            name='Seccion',
            parent=self.styles['Heading2'],
            fontSize=11,
            textColor=colors.white,
            backColor=colors.HexColor('#1F4E78'),
            spaceAfter=8,
            spaceBefore=10,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
            leftIndent=0
        ))
        
        self.styles.add(ParagraphStyle(
            name='NormalSmall',
            parent=self.styles['Normal'],
            fontSize=9,
            spaceAfter=4,
            leading=12
        ))
    
    def generar(self, datos, ruta_salida):
        """Genera archivo PDF limpio y profesional"""
        doc = SimpleDocTemplate(
            ruta_salida,
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        story = []
        
        # === ENCABEZADO ===
        story.append(Paragraph(datos.get('empresa', 'VUBA LOGISTICS'), self.styles['TituloEmpresa']))
        story.append(Paragraph('ACTA DE RECEPCIÓN Y RESPONSABILIDAD<br/>DE EQUIPOS DE CÓMPUTO', 
                              self.styles['TituloDocumento']))
        story.append(Spacer(1, 0.2*inch))
        
        # === INFORMACIÓN GENERAL ===
        tabla_info = Table([
            ['Fecha:', datos.get('fecha', ''), 'No. Responsiva:', datos.get('numero_responsiva', '')]
        ], colWidths=[1*inch, 2*inch, 1.5*inch, 2*inch])
        
        tabla_info.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(tabla_info)
        story.append(Spacer(1, 0.15*inch))
        
        # === DATOS DEL EMPLEADO ===
        story.append(Paragraph('DATOS DEL EMPLEADO', self.styles['Seccion']))
        story.append(Spacer(1, 0.05*inch))
        
        tabla_empleado = Table([
            ['Nombre:', datos.get('nombre_usuario', ''), 'Cédula:', datos.get('cedula', '')],
            ['Cargo:', datos.get('cargo', ''), 'Área:', datos.get('area', '')]
        ], colWidths=[1.2*inch, 2*inch, 1.2*inch, 2*inch])
        
        tabla_empleado.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(tabla_empleado)
        story.append(Spacer(1, 0.15*inch))
        
        # === EQUIPO PRINCIPAL ===
        story.append(Paragraph('EQUIPO PRINCIPAL ASIGNADO', self.styles['Seccion']))
        story.append(Spacer(1, 0.05*inch))
        
        tabla_equipo = Table([
            ['Marca:', datos.get('lap_marca', ''), 'Modelo:', datos.get('lap_modelo', '')],
            ['No. Serie:', datos.get('lap_serie', ''), 'Estado Físico:', datos.get('lap_estado', '')]
        ], colWidths=[1.2*inch, 2*inch, 1.2*inch, 2*inch])
        
        tabla_equipo.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(tabla_equipo)
        story.append(Spacer(1, 0.15*inch))
        
        # === ACCESORIOS ===
        if datos.get('accesorios'):
            story.append(Paragraph('ACCESORIOS ASIGNADOS', self.styles['Seccion']))
            story.append(Spacer(1, 0.05*inch))
            
            acc_data = [['Tipo', 'Marca', 'Modelo', 'Serie', 'Estado']]
            for acc in datos['accesorios']:
                acc_data.append([
                    acc.get('tipo', ''),
                    acc.get('marca', ''),
                    acc.get('modelo', ''),
                    acc.get('serie', ''),
                    acc.get('estado', '')
                ])
            
            tabla_acc = Table(acc_data, colWidths=[1.2*inch, 1.2*inch, 1.2*inch, 1.6*inch, 1*inch])
            tabla_acc.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), 'Helvetica', 8),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(tabla_acc)
            story.append(Spacer(1, 0.15*inch))
        
        # === RESPONSABILIDADES ===
        story.append(Paragraph('RESPONSABILIDADES DEL EMPLEADO', self.styles['Seccion']))
        story.append(Spacer(1, 0.08*inch))
        
        responsabilidades = [
            '1. El empleado es responsable del cuidado, conservación y buen uso del equipo asignado.',
            '2. El equipo debe ser devuelto en las mismas condiciones de recepción al finalizar la relación laboral.',
            '3. Cualquier daño o pérdida por negligencia será responsabilidad del empleado.',
            '4. El equipo es propiedad de la empresa y debe usarse únicamente para fines laborales.',
            '5. Debe reportar inmediatamente cualquier daño, mal funcionamiento o pérdida del equipo.'
        ]
        
        for resp in responsabilidades:
            story.append(Paragraph(resp, self.styles['NormalSmall']))
        
        story.append(Spacer(1, 0.25*inch))
        
        # === FIRMAS ===
        story.append(Paragraph('FIRMAS DE RESPONSABILIDAD', self.styles['Seccion']))
        story.append(Spacer(1, 0.3*inch))
        
        tabla_firmas = Table([
            [
                f"_______________________\n{datos.get('nombre_usuario', '')}\nEMPLEADO\nFecha: {datos.get('fecha', '')}",
                f"_______________________\n{datos.get('jefe_nombre', 'JEFE INMEDIATO')}\nJEFE INMEDIATO\nFecha: {datos.get('fecha', '')}",
                f"_______________________\n{datos.get('ti_nombre', 'DEPTO. TI')}\nDEPTO. TI\nFecha: {datos.get('fecha', '')}"
            ]
        ], colWidths=[2.1*inch, 2.1*inch, 2.1*inch])
        
        tabla_firmas.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 15),
        ]))
        story.append(tabla_firmas)
        
        # === NOTAS FINALES ===
        story.append(Spacer(1, 0.2*inch))
        story.append(Paragraph('• Este documento debe ser guardado como comprobante de entrega.', 
                              self.styles['NormalSmall']))
        story.append(Paragraph('• Se requieren las tres firmas para que el documento tenga validez.', 
                              self.styles['NormalSmall']))
        
        # Generar PDF
        doc.build(story)
        return ruta_salida


class AplicacionResponsivas:
    """Interfaz gráfica principal"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Responsivas - VUBA LOGISTICS")
        self.root.geometry("750x650")
        self.root.resizable(False, False)
        
        # Managers
        self.db = DatabaseManager()
        self.excel_gen = ExcelGenerator()
        if REPORTLAB_AVAILABLE:
            self.pdf_gen = PDFGenerator()
        
        # Carpeta de salida
        self.output_dir = Path("Responsivas_Generadas")
        self.output_dir.mkdir(exist_ok=True)
        
        self.datos_actual = None
        
        self._crear_interfaz()
    
    def _crear_interfaz(self):
        """Crea la interfaz gráfica"""
        # === ENCABEZADO ===
        frame_header = ttk.Frame(self.root, padding=15)
        frame_header.pack(fill="x")
        
        ttk.Label(frame_header, text="SISTEMA DE RESPONSIVAS", 
                 font=('Segoe UI', 20, 'bold')).pack()
        ttk.Label(frame_header, text="Departamento de Tecnología - VUBA LOGISTICS", 
                 font=('Segoe UI', 10)).pack()
        
        ttk.Separator(self.root, orient='horizontal').pack(fill='x', pady=10)
        
        # === BÚSQUEDA ===
        frame_busqueda = ttk.LabelFrame(self.root, text=" Buscar Empleado ", padding=15)
        frame_busqueda.pack(fill="x", padx=20, pady=10)
        
        ttk.Label(frame_busqueda, text="Nombre del Empleado:", 
                 font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=(0, 5))
        
        frame_input = ttk.Frame(frame_busqueda)
        frame_input.pack(fill="x")
        
        self.entry_nombre = ttk.Entry(frame_input, width=45, font=('Segoe UI', 11))
        self.entry_nombre.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.entry_nombre.bind('<Return>', lambda e: self.buscar_empleado())
        
        ttk.Button(frame_input, text="Buscar", command=self.buscar_empleado,
                  width=15).pack(side="left")
        
        ttk.Button(frame_input, text="Ver Todos", command=self.ver_todos_usuarios,
                  width=15).pack(side="left", padx=(5, 0))
        
        # === INFORMACIÓN ===
        frame_info = ttk.LabelFrame(self.root, text="Información del Empleado ", padding=15)
        frame_info.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Crear un frame con scrollbar
        canvas = tk.Canvas(frame_info, height=250)
        scrollbar = ttk.Scrollbar(frame_info, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        self.label_info = ttk.Label(scrollable_frame, 
                                    text="Busca un empleado para ver su información", 
                                    font=('Consolas', 10), justify='left')
        self.label_info.pack(anchor='nw', padx=10, pady=10)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # === BOTONES DE ACCIÓN ===
        frame_botones = ttk.Frame(self.root, padding=10)
        frame_botones.pack(fill="x", padx=20, pady=10)
        
        ttk.Button(frame_botones, text="Generar Excel",
                  command=self.generar_excel, width=22).pack(side="left", padx=5)
        
        if REPORTLAB_AVAILABLE:
            ttk.Button(frame_botones, text="Generar PDF",
                      command=self.generar_pdf, width=22).pack(side="left", padx=5)
        
        ttk.Button(frame_botones, text="Abrir Carpeta",
                  command=self.abrir_carpeta, width=22).pack(side="left", padx=5)
        
        # === BARRA DE ESTADO ===
        self.label_estado = ttk.Label(self.root, text="✓ Sistema listo", 
                                     relief="sunken", font=('Segoe UI', 9))
        self.label_estado.pack(fill="x", side="bottom")
    
    def buscar_empleado(self):
        """Busca un empleado en la base de datos"""
        nombre = self.entry_nombre.get().strip()
        
        if not nombre:
            messagebox.showwarning("Atención", "Ingresa el nombre del empleado")
            return
        
        try:
            self.datos_actual = self.db.buscar_usuario(nombre)
            
            if not self.datos_actual:
                messagebox.showinfo("No encontrado", 
                                   f"No se encontró ningún empleado con el nombre:\n'{nombre}'")
                self.label_info.config(text="Empleado no encontrado")
                self.datos_actual = None
                return
            
            # Mostrar información
            info_texto = self._formatear_info(self.datos_actual)
            self.label_info.config(text=info_texto)
            self.label_estado.config(text=f"✓ Empleado encontrado: {self.datos_actual['nombre_usuario']}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al buscar empleado:\n{str(e)}")
            self.label_estado.config(text="Error en la búsqueda")
    
    def _formatear_info(self, datos):
        """Formatea la información para mostrar"""
        texto = f"""
╔══════════════════════════════════════════════════════════════╗
║                    DATOS DEL EMPLEADO                        ║
╚══════════════════════════════════════════════════════════════╝

  Nombre:     {datos.get('nombre_usuario', 'N/A')}
  Cédula:     {datos.get('cedula', 'N/A')}
  Cargo:      {datos.get('cargo', 'N/A')}
  Área:       {datos.get('area', 'N/A')}
  Empresa:    {datos.get('empresa', 'N/A')}

╔══════════════════════════════════════════════════════════════╗
║                    EQUIPO ASIGNADO                           ║
╚══════════════════════════════════════════════════════════════╝

  Marca:      {datos.get('lap_marca', 'N/A')}
  Modelo:     {datos.get('lap_modelo', 'N/A')}
  Serie:      {datos.get('lap_serie', 'N/A')}
  Estado:     {datos.get('lap_estado', 'N/A')}
  Estatus:    {datos.get('sts', 'N/A')}
"""
        
        if datos.get('accesorios'):
            texto += f"""
╔══════════════════════════════════════════════════════════════╗
║                       ACCESORIOS                             ║
╚══════════════════════════════════════════════════════════════╝
"""
            for i, acc in enumerate(datos['accesorios'], 1):
                texto += f"""
  {i}. {acc.get('tipo', 'N/A')}
     Marca:  {acc.get('marca', 'N/A')}
     Modelo: {acc.get('modelo', 'N/A')}
     Serie:  {acc.get('serie', 'N/A')}
     Estado: {acc.get('estado', 'N/A')}
"""
        
        return texto
    
    def generar_excel(self):
        """Genera documento Excel"""
        if not self.datos_actual:
            messagebox.showwarning("Atención", "Primero busca un empleado")
            return
        
        try:
            self.label_estado.config(text="Generando Excel...")
            self.root.update()
            
            nombre_archivo = f"Responsiva_{self.datos_actual['nombre_usuario'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            ruta_completa = self.output_dir / nombre_archivo
            
            self.excel_gen.generar(self.datos_actual, str(ruta_completa))
            
            # Registrar en BD
            self.db.registrar_responsiva(
                self.datos_actual['numero_responsiva'],
                self.datos_actual['id_usuario'],
                'EXCEL',
                str(ruta_completa)
            )
            
            self.label_estado.config(text=f"✓ Excel generado: {nombre_archivo}")
            messagebox.showinfo("Éxito", 
                              f"✓ Excel generado exitosamente\n\n"
                              f"Archivo: {nombre_archivo}\n"
                              f"Ubicación: {self.output_dir}/")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar Excel:\n{str(e)}")
            self.label_estado.config(text="Error al generar Excel")
    
    def generar_pdf(self):
        """Genera documento PDF"""
        if not REPORTLAB_AVAILABLE:
            messagebox.showerror("Error", "ReportLab no está instalado.\nEjecuta: pip install reportlab")
            return
        
        if not self.datos_actual:
            messagebox.showwarning("Atención", "Primero busca un empleado")
            return
        
        try:
            self.label_estado.config(text="⏳ Generando PDF...")
            self.root.update()
            
            nombre_archivo = f"Responsiva_{self.datos_actual['nombre_usuario'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            ruta_completa = self.output_dir / nombre_archivo
            
            self.pdf_gen.generar(self.datos_actual, str(ruta_completa))
            
            # Registrar en BD
            self.db.registrar_responsiva(
                self.datos_actual['numero_responsiva'],
                self.datos_actual['id_usuario'],
                'PDF',
                str(ruta_completa)
            )
            
            self.label_estado.config(text=f"✓ PDF generado: {nombre_archivo}")
            messagebox.showinfo("Éxito", 
                              f"✓ PDF generado exitosamente\n\n"
                              f"Archivo: {nombre_archivo}\n"
                              f"Ubicación: {self.output_dir}/")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar PDF:\n{str(e)}")
            self.label_estado.config(text="Error al generar PDF")
    
    def ver_todos_usuarios(self):
        """Muestra ventana con lista de todos los usuarios"""
        ventana = tk.Toplevel(self.root)
        ventana.title("Todos los Empleados")
        ventana.geometry("700x400")
        
        ttk.Label(ventana, text="Lista de Empleados Registrados", 
                 font=('Segoe UI', 12, 'bold')).pack(pady=10)
        
        # Treeview
        frame_tree = ttk.Frame(ventana)
        frame_tree.pack(fill="both", expand=True, padx=10, pady=10)
        
        columnas = ('ID', 'Nombre', 'Cédula', 'Cargo', 'Área')
        tree = ttk.Treeview(frame_tree, columns=columnas, show='headings', height=15)
        
        for col in columnas:
            tree.heading(col, text=col)
            tree.column(col, width=130)
        
        scrollbar = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # Llenar datos
        usuarios = self.db.listar_usuarios()
        for usuario in usuarios:
            tree.insert('', 'end', values=usuario)
        
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Botón para seleccionar
        def seleccionar_usuario():
            seleccion = tree.selection()
            if seleccion:
                item = tree.item(seleccion[0])
                nombre = item['values'][1]
                self.entry_nombre.delete(0, tk.END)
                self.entry_nombre.insert(0, nombre)
                ventana.destroy()
                self.buscar_empleado()
        
        ttk.Button(ventana, text="Seleccionar Empleado", 
                  command=seleccionar_usuario).pack(pady=10)
    
    def abrir_carpeta(self):
        """Abre la carpeta de responsivas generadas"""
        import subprocess
        import platform
        
        ruta = str(self.output_dir.absolute())
        
        try:
            if platform.system() == 'Windows':
                os.startfile(ruta)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.Popen(['open', ruta])
            else:  # Linux
                subprocess.Popen(['xdg-open', ruta])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la carpeta:\n{str(e)}")


def main():
    """Función principal"""
    root = tk.Tk()
    app = AplicacionResponsivas(root)
    root.mainloop()


if __name__ == "__main__":
    main()
